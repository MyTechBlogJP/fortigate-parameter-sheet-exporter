#!/usr/bin/env python3
"""FortiGate REST API read-only parameter sheet exporter v1.0.

This program only issues HTTP GET requests. It exports allowlisted fields to a
sanitized JSON snapshot and a human-readable Excel workbook. It intentionally
does not save raw API responses or credentials.
"""

from __future__ import annotations

import argparse
import getpass
import json
import os
import ssl
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen


VERSION = "1.0.0"
USER_AGENT = f"fortigate-parameter-export/{VERSION}"

# Monitor/CMDB共通のレスポンス封筒から取り込む許可項目。results本体は保存しない。
ENVELOPE_ALLOWLIST = ("serial", "version", "build")


@dataclass(frozen=True)
class Endpoint:
    key: str
    label: str
    path: str
    vdom_scoped: bool = True
    required: bool = False


ENDPOINTS = [
    Endpoint("system_status", "System status", "monitor/system/status", False),
    Endpoint("vdoms", "VDOMs", "cmdb/system/vdom", False),
    Endpoint("interfaces", "Interfaces", "cmdb/system/interface", True, True),
    Endpoint("zones", "Zones", "cmdb/system/zone"),
    Endpoint("static_routes", "Configured static routes", "cmdb/router/static", True, True),
    Endpoint("policy_routes", "Policy routes", "cmdb/router/policy"),
    Endpoint("sdwan", "SD-WAN", "cmdb/system/sdwan"),
    Endpoint("addresses", "Addresses", "cmdb/firewall/address", True, True),
    Endpoint("address_groups", "Address groups", "cmdb/firewall/addrgrp", True, True),
    Endpoint("vips", "VIPs", "cmdb/firewall/vip"),
    Endpoint("vip_groups", "VIP groups", "cmdb/firewall/vipgrp"),
    Endpoint("services", "Services", "cmdb/firewall.service/custom", True, True),
    Endpoint("service_groups", "Service groups", "cmdb/firewall.service/group", True, True),
    Endpoint("firewall_policies", "Firewall policies", "cmdb/firewall/policy", True, True),
    Endpoint("dhcp_servers", "DHCP servers", "cmdb/system.dhcp/server"),
    Endpoint("system_global", "System global", "cmdb/system/global", False),
    Endpoint("system_dns", "System DNS", "cmdb/system/dns", False),
    Endpoint("system_ntp", "System NTP", "cmdb/system/ntp", False),
]


@dataclass
class FetchResult:
    endpoint: Endpoint
    url_path: str
    http_status: int | None
    api_status: str
    results: Any = field(default_factory=list)
    message: str = ""
    envelope: dict[str, Any] = field(default_factory=dict)

    @property
    def ok(self) -> bool:
        return self.http_status == 200 and self.api_status == "success"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "FortiGate REST APIから基本設定を読み取り専用で取得し、"
            "サニタイズ済みJSONとExcelパラメータシートへ出力します。"
        )
    )
    parser.add_argument(
        "--base-url",
        default=os.getenv("FGT_BASE_URL", "https://192.168.10.254"),
        help="FortiGate URL（既定: %(default)s）",
    )
    parser.add_argument(
        "--vdom",
        default="root",
        help="取得対象VDOM。複数指定はカンマ区切り（既定: root）",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        help="出力先。省略時は日時入りディレクトリを作成",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=30,
        help="APIタイムアウト秒（既定: 30）",
    )
    parser.add_argument(
        "--insecure",
        action="store_true",
        help="TLSサーバー証明書の検証を無効化（検証環境限定）",
    )
    parser.add_argument(
        "--version",
        action="version",
        version=f"%(prog)s {VERSION}",
    )
    return parser.parse_args()


def load_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(
            "エラー: openpyxlが必要です。Ubuntuで "
            "sudo apt install -y python3-openpyxl を実行してください。",
            file=sys.stderr,
        )
        raise SystemExit(2)
    return (
        Workbook,
        CellIsRule,
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
        get_column_letter,
    )


def api_get(
    base_url: str,
    endpoint: Endpoint,
    token: str,
    vdom: str,
    insecure: bool,
    timeout: int,
) -> FetchResult:
    query = {"vdom": vdom} if endpoint.vdom_scoped else {}
    path = f"/api/v2/{endpoint.path}"
    url = f"{base_url}{path}"
    if query:
        url += "?" + urlencode(query)

    request = Request(
        url,
        method="GET",
        headers={
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
            "User-Agent": USER_AGENT,
        },
    )
    context = ssl.create_default_context()
    if insecure:
        context.check_hostname = False
        context.verify_mode = ssl.CERT_NONE

    try:
        with urlopen(request, context=context, timeout=timeout) as response:
            status = response.status
            body = response.read().decode("utf-8")
    except HTTPError as exc:
        exc.read()
        return FetchResult(endpoint, path, exc.code, "http_error", [], "HTTP error")
    except URLError as exc:
        return FetchResult(endpoint, path, None, "connection_error", [], str(exc.reason))
    except TimeoutError:
        return FetchResult(endpoint, path, None, "timeout", [], f"timeout={timeout}s")

    try:
        payload = json.loads(body)
    except json.JSONDecodeError:
        return FetchResult(endpoint, path, status, "invalid_json", [], "JSONではないレスポンス")

    api_status = str(payload.get("status", "unknown"))
    results = payload.get("results", [])
    message = str(payload.get("error", payload.get("message", "")))[:300]
    envelope = {key: payload[key] for key in ENVELOPE_ALLOWLIST if key in payload}
    return FetchResult(endpoint, path, status, api_status, results, message, envelope)


def as_list(value: Any) -> list[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, dict):
        return [value]
    return []


def dict_rows(value: Any) -> list[dict[str, Any]]:
    return [item for item in as_list(value) if isinstance(item, dict)]


def first_dict(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    rows = dict_rows(value)
    return rows[0] if rows else {}


def scalar(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, list):
        return join_names(value)
    if isinstance(value, dict):
        return value.get("name", value.get("id", ""))
    return str(value)


def join_names(value: Any) -> str:
    if not isinstance(value, list):
        return "" if value is None else str(scalar(value))
    output: list[str] = []
    for item in value:
        if isinstance(item, dict):
            output.append(
                str(
                    item.get(
                        "name",
                        item.get("range", item.get("server", item.get("id", ""))),
                    )
                )
            )
        else:
            output.append(str(item))
    return ", ".join(part for part in output if part)


def member_names(value: Any) -> list[str]:
    if not isinstance(value, list):
        return []
    names: list[str] = []
    for item in value:
        if isinstance(item, dict):
            name = item.get("name")
        else:
            name = item
        if name not in (None, ""):
            names.append(str(name))
    return names


def ipmask_to_cidr(value: Any) -> str:
    text = scalar(value)
    if not isinstance(text, str) or not text.strip():
        return ""
    parts = text.replace("/", " ").split()
    if len(parts) != 2:
        return text
    ip, mask = parts
    try:
        octets = [int(x) for x in mask.split(".")]
        binary = "".join(f"{x:08b}" for x in octets)
        if "01" in binary or len(octets) != 4:
            return text
        return f"{ip}/{binary.count('1')}"
    except (TypeError, ValueError):
        return text


def mask_identifier(value: Any) -> str:
    """Keep only the leading 4 characters so the identifier stays unusable."""
    text = str(value or "")
    if len(text) <= 4:
        return "****" if text else ""
    return text[:4] + "*" * max(4, len(text) - 4)


def safe_hostname(value: Any, serial: Any) -> str:
    """Mask FortiGate default hostnames that expose a serial-like identifier."""
    hostname = str(value or "")
    serial_text = str(serial or "")
    if not hostname:
        return ""
    if serial_text and hostname.casefold() == serial_text.casefold():
        return mask_identifier(hostname)
    if hostname.upper().startswith("FGVM") and len(hostname) >= 12:
        return mask_identifier(hostname)
    return hostname


def endpoint_rows(fetches: dict[str, FetchResult], key: str) -> list[dict[str, Any]]:
    item = fetches.get(key)
    return dict_rows(item.results) if item and item.ok else []


def endpoint_envelope(fetches: dict[str, FetchResult], key: str) -> dict[str, Any]:
    item = fetches.get(key)
    return dict(item.envelope) if item and item.ok else {}


def endpoint_object(fetches: dict[str, FetchResult], key: str) -> dict[str, Any]:
    item = fetches.get(key)
    return first_dict(item.results) if item and item.ok else {}


def address_value(obj: dict[str, Any]) -> str:
    kind = str(obj.get("type", "ipmask"))
    if kind in {"ipmask", "wildcard"}:
        return ipmask_to_cidr(obj.get("subnet") or obj.get("wildcard"))
    if kind == "iprange":
        return f"{scalar(obj.get('start-ip'))} - {scalar(obj.get('end-ip'))}"
    if kind in {"fqdn", "wildcard-fqdn"}:
        return str(obj.get("fqdn") or obj.get("wildcard-fqdn") or "")
    if kind == "geography":
        return str(obj.get("country", ""))
    if kind == "dynamic":
        return str(obj.get("sdn-addr-type", "dynamic"))
    return str(obj.get("subnet") or obj.get("fqdn") or "")


def service_value(obj: dict[str, Any]) -> str:
    protocol = str(obj.get("protocol", ""))
    parts: list[str] = []
    for label, key in (("TCP", "tcp-portrange"), ("UDP", "udp-portrange"), ("SCTP", "sctp-portrange")):
        value = str(obj.get(key, "") or "")
        if value:
            parts.append(f"{label}/{value}")
    if protocol in {"ICMP", "ICMP6"}:
        icmp_type = obj.get("icmptype", "any")
        icmp_code = obj.get("icmpcode", "any")
        parts.append(f"{protocol} type={icmp_type} code={icmp_code}")
    protocol_number = obj.get("protocol-number")
    if protocol == "IP" and protocol_number not in (None, ""):
        parts.append(f"IP protocol={protocol_number}")
    return "; ".join(parts) or protocol


def normalize(fetches: dict[str, FetchResult], vdom: str, fetched_at: str, base_url: str) -> dict[str, Any]:
    validation: list[dict[str, Any]] = []

    for result in fetches.values():
        severity = "INFO" if result.ok else ("ERROR" if result.endpoint.required else "WARN")
        validation.append(
            {
                "severity": severity,
                "category": "API取得",
                "target": result.endpoint.label,
                "message": (
                    f"HTTP {result.http_status}; API status={result.api_status}"
                    + (f"; {result.message}" if result.message else "")
                ),
            }
        )

    status = endpoint_object(fetches, "system_status")
    status_meta = endpoint_envelope(fetches, "system_status")
    serial = status_meta.get("serial", "")
    system_global = endpoint_object(fetches, "system_global")
    system_dns = endpoint_object(fetches, "system_dns")
    system_ntp = endpoint_object(fetches, "system_ntp")

    hostname = safe_hostname(
        status.get("hostname") or system_global.get("hostname", ""),
        serial,
    )

    overview = [
        {"item": "Exporter Version", "value": VERSION},
        {"item": "取得日時", "value": fetched_at},
        {"item": "FortiGate URL", "value": base_url},
        {"item": "VDOM", "value": vdom},
        {"item": "Hostname", "value": hostname},
        {"item": "Model", "value": status.get("model_name") or status.get("platform_type", "")},
        {"item": "FortiOS Version", "value": status_meta.get("version", "")},
        {"item": "Build", "value": status_meta.get("build", "")},
        {"item": "Serial Number", "value": mask_identifier(serial)},
        {"item": "動作", "value": "REST API GETのみ（設定変更なし）"},
        {"item": "取得範囲", "value": "設定情報を中心に取得。Monitor APIは機器識別メタデータだけに使用。"},
        {"item": "保存方式", "value": "許可項目だけを保存。API生レスポンスは保存しない。"},
    ]

    vdom_rows = []
    for obj in endpoint_rows(fetches, "vdoms"):
        vdom_rows.append(
            {"name": obj.get("name", ""), "status": obj.get("status", ""), "short-name": obj.get("short-name", "")}
        )
    if not vdom_rows:
        vdom_rows.append({"name": vdom, "status": "", "short-name": ""})

    interfaces = []
    for obj in endpoint_rows(fetches, "interfaces"):
        mode = str(obj.get("mode", ""))
        configured_ip = "" if mode == "dhcp" else ipmask_to_cidr(obj.get("ip"))
        interfaces.append(
            {
                "name": obj.get("name", ""),
                "alias": obj.get("alias", ""),
                "type": obj.get("type", ""),
                "role": obj.get("role", ""),
                "status": obj.get("status", ""),
                "mode": mode,
                "ip": configured_ip,
                "defaultgw": obj.get("defaultgw", ""),
                "distance": obj.get("distance", ""),
                "priority": obj.get("priority", ""),
                "dns-server-override": obj.get("dns-server-override", ""),
                "allowaccess": join_names(obj.get("allowaccess")) if isinstance(obj.get("allowaccess"), list) else scalar(obj.get("allowaccess")),
                "device-identification": obj.get("device-identification", ""),
                "description": obj.get("description", ""),
            }
        )

    zones = []
    for obj in endpoint_rows(fetches, "zones"):
        zones.append(
            {
                "name": obj.get("name", ""),
                "interfaces": join_names(obj.get("interface")),
                "intrazone": obj.get("intrazone", ""),
                "description": obj.get("description", ""),
            }
        )

    static_routes = []
    for obj in endpoint_rows(fetches, "static_routes"):
        static_routes.append(
            {
                "seq-num": obj.get("seq-num", obj.get("seq_num", "")),
                "status": obj.get("status", ""),
                "dst": ipmask_to_cidr(obj.get("dst")),
                "gateway": obj.get("gateway", ""),
                "device": scalar(obj.get("device")),
                "distance": obj.get("distance", ""),
                "priority": obj.get("priority", ""),
                "comment": obj.get("comment", ""),
            }
        )

    policy_routes = []
    for obj in endpoint_rows(fetches, "policy_routes"):
        policy_routes.append(
            {
                "seq-num": obj.get("seq-num", ""),
                "status": obj.get("status", ""),
                "input-device": scalar(obj.get("input-device")),
                "src": ipmask_to_cidr(obj.get("src")),
                "dst": ipmask_to_cidr(obj.get("dst")),
                "protocol": obj.get("protocol", ""),
                "start-port": obj.get("start-port", ""),
                "end-port": obj.get("end-port", ""),
                "gateway": obj.get("gateway", ""),
                "output-device": scalar(obj.get("output-device")),
                "comments": obj.get("comments", ""),
            }
        )

    sdwan_obj = endpoint_object(fetches, "sdwan")
    sdwan_zones = []
    for obj in dict_rows(sdwan_obj.get("zone")):
        sdwan_zones.append(
            {"name": obj.get("name", ""), "service-sla-tie-break": obj.get("service-sla-tie-break", ""), "minimum-sla-meet-members": obj.get("minimum-sla-meet-members", "")}
        )
    sdwan_members = []
    for obj in dict_rows(sdwan_obj.get("members")):
        sdwan_members.append(
            {
                "seq-num": obj.get("seq-num", ""),
                "interface": scalar(obj.get("interface")),
                "zone": scalar(obj.get("zone")),
                "gateway": obj.get("gateway", ""),
                "cost": obj.get("cost", ""),
                "priority": obj.get("priority", ""),
                "status": obj.get("status", ""),
                "comment": obj.get("comment", ""),
            }
        )
    sdwan_health_checks = []
    for obj in dict_rows(sdwan_obj.get("health-check")):
        sla_rows = dict_rows(obj.get("sla"))
        sla_text = "; ".join(
            f"id={sla.get('id', '')} latency={sla.get('latency-threshold', '')}ms "
            f"jitter={sla.get('jitter-threshold', '')}ms loss={sla.get('packetloss-threshold', '')}%"
            for sla in sla_rows
        )
        sdwan_health_checks.append(
            {
                "name": obj.get("name", ""),
                "server": join_names(obj.get("server")) if isinstance(obj.get("server"), list) else scalar(obj.get("server")),
                "protocol": obj.get("protocol", ""),
                "interval": obj.get("interval", ""),
                "failtime": obj.get("failtime", ""),
                "recoverytime": obj.get("recoverytime", ""),
                "members": join_names(obj.get("members")),
                "sla": sla_text,
            }
        )
    sdwan_rules = []
    for obj in dict_rows(sdwan_obj.get("service")):
        sdwan_rules.append(
            {
                "id": obj.get("id", ""),
                "name": obj.get("name", ""),
                "status": obj.get("status", ""),
                "mode": obj.get("mode", ""),
                "src": join_names(obj.get("src")),
                "dst": join_names(obj.get("dst")),
                "internet-service": obj.get("internet-service", ""),
                "protocol": obj.get("protocol", ""),
                "start-port": obj.get("start-port", ""),
                "end-port": obj.get("end-port", ""),
                "priority-members": join_names(obj.get("priority-members")),
                "sla": join_names(obj.get("sla")),
            }
        )

    address_objects: dict[str, dict[str, Any]] = {}
    addresses = []
    for obj in endpoint_rows(fetches, "addresses"):
        name = str(obj.get("name", ""))
        address_objects[name] = obj
        addresses.append(
            {
                "name": name,
                "type": obj.get("type", ""),
                "value": address_value(obj),
                "associated-interface": scalar(obj.get("associated-interface")),
                "visibility": obj.get("visibility", ""),
                "comment": obj.get("comment", ""),
            }
        )

    address_group_objects: dict[str, dict[str, Any]] = {}
    address_groups = []
    for obj in endpoint_rows(fetches, "address_groups"):
        name = str(obj.get("name", ""))
        address_group_objects[name] = obj
        address_groups.append(
            {
                "name": name,
                "members": join_names(obj.get("member")),
                "exclude": obj.get("exclude", ""),
                "exclude-members": join_names(obj.get("exclude-member")),
                "comment": obj.get("comment", ""),
            }
        )

    vip_objects: dict[str, dict[str, Any]] = {}
    vips = []
    for obj in endpoint_rows(fetches, "vips"):
        name = str(obj.get("name", ""))
        vip_objects[name] = obj
        vips.append(
            {
                "name": name,
                "type": obj.get("type", ""),
                "extip": obj.get("extip", ""),
                "mappedip": join_names(obj.get("mappedip")),
                "extintf": scalar(obj.get("extintf")),
                "portforward": obj.get("portforward", ""),
                "protocol": obj.get("protocol", ""),
                "extport": obj.get("extport", ""),
                "mappedport": obj.get("mappedport", ""),
                "comment": obj.get("comment", ""),
            }
        )

    vip_group_objects: dict[str, dict[str, Any]] = {}
    vip_groups = []
    for obj in endpoint_rows(fetches, "vip_groups"):
        name = str(obj.get("name", ""))
        vip_group_objects[name] = obj
        vip_groups.append(
            {"name": name, "members": join_names(obj.get("member")), "interface": scalar(obj.get("interface")), "comments": obj.get("comments", "")}
        )

    service_objects: dict[str, dict[str, Any]] = {}
    services = []
    for obj in endpoint_rows(fetches, "services"):
        name = str(obj.get("name", ""))
        service_objects[name] = obj
        services.append(
            {
                "name": name,
                "protocol": obj.get("protocol", ""),
                "ports-or-type": service_value(obj),
                "category": scalar(obj.get("category")),
                "visibility": obj.get("visibility", ""),
                "comment": obj.get("comment", ""),
            }
        )

    service_group_objects: dict[str, dict[str, Any]] = {}
    service_groups = []
    for obj in endpoint_rows(fetches, "service_groups"):
        name = str(obj.get("name", ""))
        service_group_objects[name] = obj
        service_groups.append(
            {"name": name, "members": join_names(obj.get("member")), "comment": obj.get("comment", "")}
        )

    zone_map = {row["name"]: row["interfaces"] for row in zones if row.get("name")}
    sdwan_zone_map: dict[str, list[str]] = {}
    for row in sdwan_members:
        zone_name = str(row.get("zone") or "virtual-wan-link")
        sdwan_zone_map.setdefault(zone_name, []).append(str(row.get("interface", "")))

    def resolve_address(name: str, trail: tuple[str, ...] = ()) -> str:
        if name in {"all", "all_ipv4"}:
            return "0.0.0.0/0"
        if name == "all_ipv6":
            return "::/0"
        if name in trail:
            return f"循環参照({name})"
        if name in address_objects:
            return address_value(address_objects[name])
        if name in vip_objects:
            vip = vip_objects[name]
            mapped = join_names(vip.get("mappedip"))
            return f"VIP {vip.get('extip', '')} -> {mapped}"
        vip_group = vip_group_objects.get(name)
        if vip_group:
            return "(" + " | ".join(
                f"{member}={resolve_address(member, trail + (name,))}"
                for member in member_names(vip_group.get("member"))
            ) + ")"
        group = address_group_objects.get(name)
        if group:
            return "(" + " | ".join(
                f"{member}={resolve_address(member, trail + (name,))}"
                for member in member_names(group.get("member"))
            ) + ")"
        validation.append({"severity": "WARN", "category": "参照解決", "target": name, "message": "Address/VIPオブジェクトを解決できません"})
        return f"未解決({name})"

    def resolve_service(name: str, trail: tuple[str, ...] = ()) -> str:
        if name == "ALL":
            return "All protocols/services"
        if name in trail:
            return f"循環参照({name})"
        if name in service_objects:
            return service_value(service_objects[name])
        group = service_group_objects.get(name)
        if group:
            return "(" + " | ".join(
                f"{member}={resolve_service(member, trail + (name,))}"
                for member in member_names(group.get("member"))
            ) + ")"
        validation.append({"severity": "WARN", "category": "参照解決", "target": name, "message": "Serviceオブジェクトを解決できません"})
        return f"未解決({name})"

    def resolve_interface(name: str) -> str:
        if name in zone_map:
            return f"Zone {name}: {zone_map[name]}"
        if name in sdwan_zone_map:
            members = ", ".join(x for x in sdwan_zone_map[name] if x)
            return f"SD-WAN Zone {name}: {members}"
        return name

    firewall_policies = []
    policy_resolved = []
    for obj in endpoint_rows(fetches, "firewall_policies"):
        src_names = member_names(obj.get("srcaddr"))
        dst_names = member_names(obj.get("dstaddr"))
        svc_names = member_names(obj.get("service"))
        in_names = member_names(obj.get("srcintf"))
        out_names = member_names(obj.get("dstintf"))
        policy_id = obj.get("policyid", "")
        policy_name = obj.get("name", "")
        firewall_policies.append(
            {
                "policyid": policy_id,
                "name": policy_name,
                "status": obj.get("status", ""),
                "srcintf": ", ".join(in_names),
                "dstintf": ", ".join(out_names),
                "srcaddr": ", ".join(src_names),
                "dstaddr": ", ".join(dst_names),
                "action": obj.get("action", ""),
                "schedule": obj.get("schedule", ""),
                "service": ", ".join(svc_names),
                "nat": obj.get("nat", ""),
                "ippool": obj.get("ippool", ""),
                "poolname": join_names(obj.get("poolname")),
                "logtraffic": obj.get("logtraffic", ""),
                "utm-status": obj.get("utm-status", ""),
                "comments": obj.get("comments", ""),
            }
        )
        policy_resolved.append(
            {
                "policyid": policy_id,
                "name": policy_name,
                "incoming": "; ".join(resolve_interface(x) for x in in_names),
                "outgoing-configured": "; ".join(resolve_interface(x) for x in out_names),
                "source-resolved": "; ".join(f"{x}={resolve_address(x)}" for x in src_names),
                "destination-resolved": "; ".join(f"{x}={resolve_address(x)}" for x in dst_names),
                "service-resolved": "; ".join(f"{x}={resolve_service(x)}" for x in svc_names),
                "action": obj.get("action", ""),
                "nat": obj.get("nat", ""),
                "note": "実際の転送経路はルーティング、Policy Route、SD-WAN Ruleも併せて確認",
            }
        )
        if "all" in src_names:
            validation.append({"severity": "INFO", "category": "Policy確認", "target": f"Policy {policy_id}", "message": "送信元がallです"})
        if "all" in dst_names:
            validation.append({"severity": "INFO", "category": "Policy確認", "target": f"Policy {policy_id}", "message": "宛先がallです"})
        if "ALL" in svc_names:
            validation.append({"severity": "INFO", "category": "Policy確認", "target": f"Policy {policy_id}", "message": "ServiceがALLです"})

    dhcp_servers = []
    for obj in endpoint_rows(fetches, "dhcp_servers"):
        ranges = []
        for iprange in dict_rows(obj.get("ip-range")):
            ranges.append(f"{iprange.get('start-ip', '')} - {iprange.get('end-ip', '')}")
        dhcp_servers.append(
            {
                "id": obj.get("id", ""),
                "status": obj.get("status", ""),
                "interface": scalar(obj.get("interface")),
                "default-gateway": obj.get("default-gateway", ""),
                "netmask": obj.get("netmask", ""),
                "ip-range": "; ".join(ranges),
                "dns-service": obj.get("dns-service", ""),
                "dns-server1": obj.get("dns-server1", ""),
                "dns-server2": obj.get("dns-server2", ""),
                "lease-time": obj.get("lease-time", ""),
                "domain": obj.get("domain", ""),
            }
        )

    operation_mode = str(status.get("operation_mode", status.get("opmode", "")) or "")

    ntp_servers = []
    for obj in dict_rows(system_ntp.get("ntpserver")):
        ntp_servers.append(str(obj.get("server", "")))
    system_settings = [
        {"section": "System", "item": "Hostname", "value": hostname},
        {"section": "System", "item": "Timezone", "value": system_global.get("timezone", "")},
        {"section": "System", "item": "Operation Mode", "value": operation_mode or "未取得（対応レスポンスキー未確認）"},
        {"section": "DNS", "item": "Primary", "value": system_dns.get("primary", "")},
        {"section": "DNS", "item": "Secondary", "value": system_dns.get("secondary", "")},
        {"section": "DNS", "item": "Protocol", "value": system_dns.get("protocol", "")},
        {"section": "NTP", "item": "Status", "value": system_ntp.get("ntpsync", "")},
        {"section": "NTP", "item": "Type", "value": system_ntp.get("type", "")},
        {"section": "NTP", "item": "Servers", "value": ", ".join(ntp_servers)},
    ]

    return {
        "metadata": {
            "exporter_version": VERSION,
            "fetched_at": fetched_at,
            "vdom": vdom,
            "read_only": True,
            "raw_api_responses_saved": False,
        },
        "Overview": overview,
        "VDOMs": vdom_rows,
        "Interfaces": interfaces,
        "Zones": zones,
        "Configured_Static_Routes": static_routes,
        "Policy_Routes": policy_routes,
        "SDWAN_Zones": sdwan_zones,
        "SDWAN_Members": sdwan_members,
        "SDWAN_Health_Checks": sdwan_health_checks,
        "SDWAN_Rules": sdwan_rules,
        "Addresses": addresses,
        "Address_Groups": address_groups,
        "VIPs": vips,
        "VIP_Groups": vip_groups,
        "Services": services,
        "Service_Groups": service_groups,
        "Firewall_Policies": firewall_policies,
        "Policy_Resolved": policy_resolved,
        "DHCP_Servers": dhcp_servers,
        "System_Settings": system_settings,
        "Validation": validation,
    }


SHEET_SPECS: dict[str, tuple[str, list[tuple[str, str, int]]]] = {
    "Overview": ("取得概要", [("item", "項目", 25), ("value", "値", 70)]),
    "VDOMs": ("VDOM", [("name", "VDOM", 24), ("status", "Status", 14), ("short-name", "Short Name", 24)]),
    "Interfaces": ("インターフェース", [("name", "Name", 20), ("alias", "Alias", 22), ("type", "Type", 15), ("role", "Role", 12), ("status", "Administrative Status", 20), ("mode", "Addressing Mode", 16), ("ip", "Configured IP Address", 24), ("defaultgw", "Default Gateway Acquisition", 24), ("distance", "Distance", 12), ("priority", "Priority", 12), ("dns-server-override", "DNS Server Override", 20), ("allowaccess", "Management Access", 26), ("device-identification", "Device Identification", 20), ("description", "Description", 36)]),
    "Zones": ("Zone", [("name", "Zone", 24), ("interfaces", "Member Interfaces", 42), ("intrazone", "Intrazone", 14), ("description", "Description", 36)]),
    "Configured_Static_Routes": ("Configured Static Route", [("seq-num", "Seq", 10), ("status", "Status", 12), ("dst", "Destination", 22), ("gateway", "Gateway", 20), ("device", "Interface / Zone", 22), ("distance", "Distance", 12), ("priority", "Priority", 12), ("comment", "Comment", 36)]),
    "Policy_Routes": ("Policy Route", [("seq-num", "Seq", 10), ("status", "Status", 12), ("input-device", "Input", 18), ("src", "Source", 22), ("dst", "Destination", 22), ("protocol", "Protocol", 12), ("start-port", "Start Port", 12), ("end-port", "End Port", 12), ("gateway", "Gateway", 20), ("output-device", "Output", 18), ("comments", "Comments", 36)]),
    "SDWAN_Zones": ("SD-WAN Zone", [("name", "Zone", 24), ("service-sla-tie-break", "SLA Tie Break", 22), ("minimum-sla-meet-members", "Minimum SLA Members", 24)]),
    "SDWAN_Members": ("SD-WAN Member", [("seq-num", "Seq", 10), ("interface", "Interface", 20), ("zone", "Zone", 22), ("gateway", "Gateway", 20), ("cost", "Cost", 10), ("priority", "Priority", 12), ("status", "Status", 12), ("comment", "Comment", 36)]),
    "SDWAN_Health_Checks": ("SD-WAN Performance SLA", [("name", "Name", 26), ("server", "Server", 32), ("protocol", "Protocol", 14), ("interval", "Interval", 12), ("failtime", "Fail Time", 12), ("recoverytime", "Recovery Time", 15), ("members", "Members", 24), ("sla", "SLA Thresholds", 55)]),
    "SDWAN_Rules": ("SD-WAN Rule", [("id", "ID", 10), ("name", "Name", 26), ("status", "Status", 12), ("mode", "Mode", 18), ("src", "Source", 30), ("dst", "Destination", 30), ("internet-service", "Internet Service", 18), ("protocol", "Protocol", 12), ("start-port", "Start Port", 12), ("end-port", "End Port", 12), ("priority-members", "Priority Members", 24), ("sla", "SLA", 24)]),
    "Addresses": ("Firewall Address", [("name", "Name", 32), ("type", "Type", 16), ("value", "IP / FQDN / Range", 38), ("associated-interface", "Associated Interface", 22), ("visibility", "Visibility", 12), ("comment", "Comment", 42)]),
    "Address_Groups": ("Address Group", [("name", "Name", 32), ("members", "Members", 55), ("exclude", "Exclude", 12), ("exclude-members", "Exclude Members", 42), ("comment", "Comment", 42)]),
    "VIPs": ("Virtual IP", [("name", "Name", 30), ("type", "Type", 16), ("extip", "External IP", 22), ("mappedip", "Mapped IP", 28), ("extintf", "External Interface", 20), ("portforward", "Port Forward", 14), ("protocol", "Protocol", 12), ("extport", "External Port", 16), ("mappedport", "Mapped Port", 16), ("comment", "Comment", 38)]),
    "VIP_Groups": ("VIP Group", [("name", "Name", 30), ("members", "Members", 55), ("interface", "Interface", 20), ("comments", "Comments", 38)]),
    "Services": ("Firewall Service", [("name", "Name", 30), ("protocol", "Protocol", 16), ("ports-or-type", "Port / Type", 44), ("category", "Category", 24), ("visibility", "Visibility", 12), ("comment", "Comment", 38)]),
    "Service_Groups": ("Service Group", [("name", "Name", 30), ("members", "Members", 60), ("comment", "Comment", 38)]),
    "Firewall_Policies": ("Firewall Policy", [("policyid", "Policy ID", 12), ("name", "Name", 30), ("status", "Status", 12), ("srcintf", "Incoming", 20), ("dstintf", "Outgoing", 20), ("srcaddr", "Source", 34), ("dstaddr", "Destination", 34), ("action", "Action", 12), ("schedule", "Schedule", 18), ("service", "Service", 34), ("nat", "NAT", 10), ("ippool", "IP Pool", 12), ("poolname", "Pool Name", 28), ("logtraffic", "Log", 12), ("utm-status", "UTM", 12), ("comments", "Comments", 42)]),
    "Policy_Resolved": ("Policy参照オブジェクトの実値展開", [("policyid", "Policy ID", 12), ("name", "Name", 28), ("incoming", "Incoming (Resolved)", 34), ("outgoing-configured", "Outgoing (Configured)", 38), ("source-resolved", "Source (Resolved)", 60), ("destination-resolved", "Destination (Resolved)", 60), ("service-resolved", "Service (Resolved)", 60), ("action", "Action", 12), ("nat", "NAT", 10), ("note", "確認上の注意", 56)]),
    "DHCP_Servers": ("DHCP Server", [("id", "ID", 10), ("status", "Status", 12), ("interface", "Interface", 20), ("default-gateway", "Default Gateway", 20), ("netmask", "Netmask", 20), ("ip-range", "IP Range", 38), ("dns-service", "DNS Service", 18), ("dns-server1", "DNS Server 1", 20), ("dns-server2", "DNS Server 2", 20), ("lease-time", "Lease Time", 14), ("domain", "Domain", 28)]),
    "System_Settings": ("System基本設定", [("section", "Section", 18), ("item", "Item", 28), ("value", "Value", 55)]),
    "Validation": ("取得・参照検査結果", [("severity", "Severity", 12), ("category", "Category", 20), ("target", "Target", 30), ("message", "Message", 85)]),
}


def create_workbook(data: dict[str, Any], output_path: Path) -> None:
    (
        Workbook,
        CellIsRule,
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
        get_column_letter,
    ) = load_openpyxl()

    wb = Workbook()
    wb.remove(wb.active)

    dark_green = "1F6D42"
    medium_green = "2E8B57"
    light_green = "EAF4EE"
    very_light = "F7FAF8"
    white = "FFFFFF"
    gray = "667085"
    amber = "FFF3CD"
    red = "F8D7DA"
    thin_gray = Side(style="thin", color="D0D5DD")

    multiple_vdoms = "," in str(data["metadata"]["vdom"])
    for sheet_name, (title, base_columns) in SHEET_SPECS.items():
        columns = list(base_columns)
        if multiple_vdoms and sheet_name not in {"Overview", "VDOMs", "System_Settings"}:
            columns.insert(0, ("vdom", "VDOM", 18))
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A4"

        last_col = get_column_letter(len(columns))
        ws.merge_cells(f"A1:{last_col}1")
        title_cell = ws["A1"]
        title_cell.value = f"FortiGate Parameter Sheet | {title}"
        title_cell.font = Font(size=15, bold=True, color=white)
        title_cell.fill = PatternFill("solid", fgColor=dark_green)
        title_cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[1].height = 28

        rows = data.get(sheet_name, [])
        ws.merge_cells(f"A2:{last_col}2")
        note = ws["A2"]
        note.value = f"VDOM: {data['metadata']['vdom']}  |  取得日時: {data['metadata']['fetched_at']}  |  件数: {len(rows)}"
        note.font = Font(size=9, color=gray)
        note.fill = PatternFill("solid", fgColor=very_light)
        note.alignment = Alignment(vertical="center")

        for col_index, (_, header, width) in enumerate(columns, start=1):
            cell = ws.cell(3, col_index, header)
            cell.font = Font(size=10, bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=medium_green)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin_gray)
            ws.column_dimensions[get_column_letter(col_index)].width = width
        ws.row_dimensions[3].height = 32

        if not rows:
            ws.merge_cells(f"A4:{last_col}4")
            empty = ws["A4"]
            empty.value = "設定なし、または取得対象なし"
            empty.font = Font(italic=True, color=gray)
            empty.fill = PatternFill("solid", fgColor=very_light)
            empty.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[4].height = 26
            filter_end_row = 3
        else:
            for row_index, row in enumerate(rows, start=4):
                for col_index, (key, _, _) in enumerate(columns, start=1):
                    value = row.get(key, "")
                    if isinstance(value, (dict, list)):
                        value = json.dumps(value, ensure_ascii=False)
                    cell = ws.cell(row_index, col_index, value)
                    cell.font = Font(size=10)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = Border(bottom=thin_gray)
                    if row_index % 2 == 0:
                        cell.fill = PatternFill("solid", fgColor=light_green)
                ws.row_dimensions[row_index].height = 31
            filter_end_row = 3 + len(rows)

        ws.auto_filter.ref = f"A3:{last_col}{filter_end_row}"
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_title_rows = "1:3"
        ws.auto_filter.ref = f"A3:{last_col}{filter_end_row}"

        if sheet_name == "Validation" and rows:
            severity_range = f"A4:A{3 + len(rows)}"
            ws.conditional_formatting.add(
                severity_range,
                CellIsRule(operator="equal", formula=['"WARN"'], fill=PatternFill("solid", fgColor=amber)),
            )
            ws.conditional_formatting.add(
                severity_range,
                CellIsRule(operator="equal", formula=['"ERROR"'], fill=PatternFill("solid", fgColor=red)),
            )

    wb.save(output_path)


def fetch_vdom(
    base_url: str,
    token: str,
    vdom: str,
    insecure: bool,
    timeout: int,
) -> dict[str, FetchResult]:
    fetches: dict[str, FetchResult] = {}
    for endpoint in ENDPOINTS:
        fetches[endpoint.key] = api_get(base_url, endpoint, token, vdom, insecure, timeout)
    return fetches


def merge_vdom_data(items: list[dict[str, Any]]) -> dict[str, Any]:
    if len(items) == 1:
        return items[0]

    merged: dict[str, Any] = {
        "metadata": {
            "exporter_version": VERSION,
            "fetched_at": items[0]["metadata"]["fetched_at"],
            "vdom": ",".join(item["metadata"]["vdom"] for item in items),
            "read_only": True,
            "raw_api_responses_saved": False,
        }
    }
    for sheet_name in SHEET_SPECS:
        merged_rows: list[dict[str, Any]] = []
        for item in items:
            vdom = item["metadata"]["vdom"]
            for row in item.get(sheet_name, []):
                copied = dict(row)
                if sheet_name not in {"Overview", "VDOMs", "System_Settings"}:
                    copied = {"vdom": vdom, **copied}
                merged_rows.append(copied)
        merged[sheet_name] = merged_rows
    return merged


def main() -> int:
    args = parse_args()
    base_url = args.base_url.rstrip("/")
    if not base_url.startswith("https://"):
        print("エラー: --base-urlはhttps://で指定してください", file=sys.stderr)
        return 2
    if args.timeout < 1:
        print("エラー: --timeoutは1以上を指定してください", file=sys.stderr)
        return 2

    vdoms = [item.strip() for item in args.vdom.split(",") if item.strip()]
    if not vdoms:
        print("エラー: --vdomが空です", file=sys.stderr)
        return 2

    token = getpass.getpass("FortiGate API token: ")
    if not token:
        print("エラー: APIトークンが空です", file=sys.stderr)
        return 2

    timestamp = datetime.now().astimezone().strftime("%Y%m%d_%H%M%S")
    output_dir = args.output_dir or Path(f"fortigate_parameter_export_{timestamp}")
    if output_dir.exists():
        print(f"エラー: 出力先がすでに存在します: {output_dir}", file=sys.stderr)
        return 2
    output_dir.mkdir(parents=True, exist_ok=False)
    output_dir.chmod(0o700)
    fetched_at = datetime.now().astimezone().isoformat(timespec="seconds")

    try:
        normalized_items = []
        all_fetches: list[dict[str, FetchResult]] = []
        for vdom in vdoms:
            fetches = fetch_vdom(base_url, token, vdom, args.insecure, args.timeout)
            all_fetches.append(fetches)
            normalized_items.append(normalize(fetches, vdom, fetched_at, base_url))

        data = merge_vdom_data(normalized_items)
        snapshot_path = output_dir / "fortigate_sanitized_snapshot.json"
        snapshot_path.write_text(
            json.dumps(data, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        snapshot_path.chmod(0o600)

        excel_path = output_dir / "fortigate_parameter_sheet.xlsx"
        create_workbook(data, excel_path)
        excel_path.chmod(0o600)

        required_failures = [
            result
            for fetches in all_fetches
            for result in fetches.values()
            if result.endpoint.required and not result.ok
        ]
        optional_failures = [
            result
            for fetches in all_fetches
            for result in fetches.values()
            if not result.endpoint.required and not result.ok
        ]
    except Exception as exc:
        print(f"エラー: {exc}", file=sys.stderr)
        return 1
    finally:
        token = ""

    print("読み取り専用エクスポートが完了しました。")
    print(f"JSON : {snapshot_path}")
    print(f"Excel: {excel_path}")
    print("API生レスポンスと資格情報は保存していません。")
    print("FortiGateの設定変更は行っていません。")
    if optional_failures:
        print(f"注意: 任意項目で{len(optional_failures)}件の取得失敗があります。Validationシートを確認してください。")
    if required_failures:
        print(f"警告: 必須項目で{len(required_failures)}件の取得失敗があります。成果物は部分取得です。", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
